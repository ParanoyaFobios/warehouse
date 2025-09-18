from warehouse1.models import MaterialOperation
from warehouse2.models import ProductOperation
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
import io
from openpyxl.styles import PatternFill
from django.utils.dateparse import parse_datetime
from django.utils import timezone


def get_unified_movement_data(filters):
    """
    Собирает данные из обоих журналов операций,
    стандартизирует и возвращает единый отсортированный список.
    """
    start_date = filters.get('start_date')
    end_date = filters.get('end_date')
    operation_type = filters.get('operation_type')
    item_search = filters.get('item_search')

    # 1. Получаем операции с продукцией (Склад 2)
    product_ops_qs = ProductOperation.objects.select_related('product', 'user').all()
    if start_date:
        product_ops_qs = product_ops_qs.filter(timestamp__gte=start_date)
    if end_date:
        product_ops_qs = product_ops_qs.filter(timestamp__lte=end_date)
    if operation_type:
        product_ops_qs = product_ops_qs.filter(operation_type=operation_type)
    if item_search:
        product_ops_qs = product_ops_qs.filter(
            Q(product__name__icontains=item_search) | Q(product__sku__icontains=item_search)
        )
    
    # 2. Получаем операции с материалами (Склад 1)
    material_ops_qs = MaterialOperation.objects.select_related('material', 'user', 'material__unit').all()
    if start_date:
        material_ops_qs = material_ops_qs.filter(date__gte=start_date)
    if end_date:
        material_ops_qs = material_ops_qs.filter(date__lte=end_date)
    # Адаптируем типы операций для материалов
    if operation_type in ['incoming', 'outgoing', 'adjustment']:
        material_ops_qs = material_ops_qs.filter(operation_type=operation_type)
    elif operation_type: # Если выбран тип, которого нет у материалов
        material_ops_qs = material_ops_qs.none()
    if item_search:
        material_ops_qs = material_ops_qs.filter(
            Q(material__name__icontains=item_search) | Q(material__article__icontains=item_search)
        )

    # 3. Стандартизируем и объединяем
    unified_list = []
    for op in product_ops_qs:
        unified_list.append({
            'timestamp': op.timestamp,
            'item_name': op.product.name,
            'item_sku': op.product.sku,
            'warehouse': 'Готовая продукция',
            'operation': op.get_operation_type_display(),
            'quantity': op.quantity,
            'user': op.user.username if op.user else 'N/A',
            'source': str(op.source) if op.source else 'Ручная операция'
        })
    
    for op in material_ops_qs:
        unified_list.append({
            'timestamp': op.date,
            'item_name': op.material.name,
            'item_sku': op.material.article,
            'warehouse': 'Сырье и материалы',
            'operation': op.get_operation_type_display(),
            'quantity': op.quantity,
            'user': op.user.username if op.user else 'N/A',
            'source': op.outgoing_category.name if op.outgoing_category else 'Приемка'
        })

    # 4. Сортируем общий список по дате
    unified_list.sort(key=lambda x: x['timestamp'], reverse=True)
    
    return unified_list


def generate_movement_report_excel(operations):
    """
    Принимает список операций и генерирует HttpResponse с Excel-файлом.
    """
    # Создаем книгу и лист в памяти
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Движение товаров'

    # 1. Заголовки с форматированием
    headers = [
        "Дата и время", "Склад", "Товар/Материал", "Артикул", 
        "Операция", "Количество", "Пользователь", "Основание/Комментарий"
    ]
    
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_title)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # 2. Заполняем данными
    # 2. Заполняем данными
    for row_num, op in enumerate(operations, 2):
        # <<< КЛЮЧЕВОЕ ИСПРАВЛЕНИЕ ЗДЕСЬ >>>
        # Получаем объект datetime из словаря
        dt_object = op['timestamp']
        
        # Делаем его "наивным" (без часового пояса)
        # timezone.localtime() сначала переводит в локальное время, 
        # а .replace(tzinfo=None) затем убирает информацию о таймзоне.
        naive_datetime = timezone.localtime(dt_object).replace(tzinfo=None)
        
        # Записываем уже "наивный" объект в ячейку
        sheet.cell(row=row_num, column=1, value=naive_datetime)
        
        sheet.cell(row=row_num, column=2, value=op['warehouse'])
        sheet.cell(row=row_num, column=3, value=op['item_name'])
        sheet.cell(row=row_num, column=4, value=op['item_sku'])
        sheet.cell(row=row_num, column=5, value=op['operation'])
        sheet.cell(row=row_num, column=6, value=op['quantity'])
        sheet.cell(row=row_num, column=7, value=op['user'])
        sheet.cell(row=row_num, column=8, value=op['source'])

    # 3. Настраиваем ширину колонок
    column_widths = [20, 15, 30, 15, 15, 12, 15, 30]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 4. Форматируем даты
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'DD.MM.YYYY HH:MM'

    # 5. Замораживаем первую строку (заголовки)
    sheet.freeze_panes = 'A2'

    # 6. Сохраняем файл в буфер
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # 7. Создаем HttpResponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="movement_report.xlsx"'
    
    return response