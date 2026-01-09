import pytest
from unittest.mock import patch
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta, date
import json
from warehouse1.models import Material
from warehouse2.models import ShipmentItem, Product
import io
from openpyxl import load_workbook
from decimal import Decimal
from warehouse1.models import MaterialOperation
from warehouse2.models import ProductOperation
from django.contrib.contenttypes.models import ContentType


# ==================== ReportsHomeView ====================
@pytest.mark.django_db
class TestReportsHomeView:
    def test_reports_home_view_renders_correct_template(self, client, user):
        """Тест, что ReportsHomeView использует правильный шаблон"""
        client.force_login(user)
        url = reverse('reports_home')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'reports/reports_home.html' in [t.name for t in response.templates]
    
    def test_reports_home_view_context(self, client, user):
        """Тест контекста ReportsHomeView"""
        client.force_login(user)
        url = reverse('reports_home')
        response = client.get(url)
        
        assert response.context['page_title'] == 'Reports Home'
        assert response.context['user'] == user


# ==================== SalesOverTimeView ====================
@pytest.mark.django_db
class TestSalesOverTimeView:
    def test_sales_over_time_view_renders_correct_template(self, client, user):
        """Тест, что SalesOverTimeView использует правильный шаблон"""
        client.force_login(user)
        url = reverse('sales_over_time')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'reports/sales_over_time.html' in [t.name for t in response.templates]


# ==================== sales_chart_data_api ====================
@pytest.mark.django_db
class TestSalesChartDataApi:
    def test_sales_chart_data_api_no_data(self, client, user):
        """Тест API данных графика продаж без данных"""
        client.force_login(user)
        url = reverse('sales_chart_data_api')
        response = client.get(url)
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data
        assert len(data['labels']) == 0
        assert len(data['data']) == 0
    
    def test_sales_chart_data_api_with_period_week(self, client, user, shipment, product):
        """Тест API с периодом неделя и данными"""
        # Создаем отгруженные Shipment с датами
        shipment.status = 'shipped'
        shipment.shipped_at = timezone.now() - timedelta(days=3)
        shipment.save()
        
        # Создаем ShipmentItem
        ShipmentItem.objects.create(
            shipment=shipment,
            product=product,
            quantity=5,
            price=1000.00
        )
        
        client.force_login(user)
        url = reverse('sales_chart_data_api')
        response = client.get(url, {'period': 'week'})
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data
    
    def test_sales_chart_data_api_with_period_month(self, client, user, shipment, product):
        """Тест API с периодом месяц"""
        client.force_login(user)
        url = reverse('sales_chart_data_api')
        response = client.get(url, {'period': 'month'})
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data
    
    def test_sales_chart_data_api_with_period_year(self, client, user, shipment, product):
        """Тест API с периодом год"""
        client.force_login(user)
        url = reverse('sales_chart_data_api')
        response = client.get(url, {'period': 'year'})
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data


# ==================== MovementReportView ====================
@pytest.mark.django_db
class TestMovementReportView:
    def test_movement_report_view_renders_correct_template(self, client, user):
        """Тест, что MovementReportView использует правильный шаблон"""
        client.force_login(user)
        url = reverse('movement_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'reports/movement_report.html' in [t.name for t in response.templates]
    
    def test_movement_report_view_with_filters(self, client, user, material_operation_incoming, product_operation_incoming):
        """Тест MovementReportView с фильтрами"""
        client.force_login(user)
        url = reverse('movement_report')
        
        # Тестируем разные фильтры
        test_cases = [
            {},  # Без фильтров
            {'operation_type': 'incoming'},
            {'start_date': date.today() - timedelta(days=7)},
            {'end_date': date.today()},
            {'item_search': 'Тест'}
        ]
        
        for params in test_cases:
            response = client.get(url, params)
            assert response.status_code == 200
            assert 'filter_form' in response.context
            assert 'page_obj' in response.context
    
    def test_movement_report_view_pagination(self, client, user):
        """Тест пагинации в MovementReportView"""
        client.force_login(user)
        url = reverse('movement_report')
        
        # Тестируем разные значения per_page
        per_page_options = [25, 50, 100]
        
        for per_page in per_page_options:
            response = client.get(url, {'per_page': per_page})
            assert response.status_code == 200
            assert response.context['per_page'] == per_page
    
    def test_movement_report_view_export_excel(self, client, user, material_operation_incoming, product_operation_incoming):
        """Тест экспорта в Excel"""
        client.force_login(user)
        url = reverse('movement_report')
        
        response = client.get(url, {'export_excel': 'true'})
        
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'attachment; filename="movement_report.xlsx"' in response['Content-Disposition']
        
        # Проверяем содержимое Excel файла
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        
        # Проверяем заголовки
        headers = [
            "Дата и время", "Склад", "Товар/Материал", "Артикул", 
            "Операция", "Количество", "Пользователь", "Основание/Комментарий"
        ]
        
        for col, header in enumerate(headers, start=1):
            assert sheet.cell(row=1, column=col).value == header
        
        # Проверяем, что есть данные (минимум 2 строки, включая заголовок)
        assert sheet.max_row >= 2


# ==================== SalesByProductReportView ====================
@pytest.mark.django_db
class TestSalesByProductReportView:
    def test_sales_by_product_report_view_renders_correct_template(self, client, user):
        """Тест, что SalesByProductReportView использует правильный шаблон"""
        client.force_login(user)
        url = reverse('sales_by_product_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'reports/sales_by_product_report.html' in [t.name for t in response.templates]
    
    def test_sales_by_product_report_view_context(self, client, user):
        """Тест контекста SalesByProductReportView"""
        client.force_login(user)
        url = reverse('sales_by_product_report')
        response = client.get(url)
        
        assert 'form' in response.context


# ==================== sales_by_product_api ====================
@pytest.mark.django_db
class TestSalesByProductApi:
    def test_sales_by_product_api_no_data(self, client, user):
        """Тест API топ-10 товаров без данных"""
        client.force_login(user)
        url = reverse('sales_by_product_api')
        response = client.get(url)
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data
        assert len(data['labels']) == 0
        assert len(data['data']) == 0
    
    def test_sales_by_product_api_with_data(self, client, user, product, shipment):
        """Тест API топ-10 товаров с данными"""
        # Создаем отгруженный Shipment
        shipment.status = 'shipped'
        shipment.shipped_at = timezone.now()
        shipment.save()
        
        # Создаем ShipmentItem
        ShipmentItem.objects.create(
            shipment=shipment,
            product=product,
            quantity=10,
            price=1000.00
        )
        
        client.force_login(user)
        url = reverse('sales_by_product_api')
        
        # Тест с датами
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()
        
        response = client.get(url, {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d')
        })
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data
        assert len(data['labels']) > 0
        assert len(data['data']) > 0
    
    def test_sales_by_product_api_invalid_date_format(self, client, user):
        """Тест API с невалидным форматом даты"""
        client.force_login(user)
        url = reverse('sales_by_product_api')
        response = client.get(url, {'start_date': 'invalid-date'})
        
        assert response.status_code == 400
        data = json.loads(response.content)
        assert 'error' in data


# ==================== sales_by_category_api ====================
@pytest.mark.django_db
class TestSalesByCategoryApi:
    def test_sales_by_category_api_no_data(self, client, user):
        """Тест API выручки по категориям без данных"""
        client.force_login(user)
        url = reverse('sales_by_category_api')
        response = client.get(url)
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data
        assert len(data['labels']) == 0
        assert len(data['data']) == 0
    
    def test_sales_by_category_api_with_data(self, client, user, product_category, product, shipment):
        """Тест API выручки по категориям с данными"""
        # Создаем продукт с категорией
        product.category = product_category
        product.save()
        
        # Создаем отгруженный Shipment
        shipment.status = 'shipped'
        shipment.shipped_at = timezone.now()
        shipment.save()
        
        # Создаем ShipmentItem
        ShipmentItem.objects.create(
            shipment=shipment,
            product=product,
            quantity=5,
            price=2000.00
        )
        
        client.force_login(user)
        url = reverse('sales_by_category_api')
        
        response = client.get(url, {
            'start_date': (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'),
            'end_date': date.today().strftime('%Y-%m-%d')
        })
        
        assert response.status_code == 200
        data = json.loads(response.content)
        assert 'labels' in data
        assert 'data' in data
        # Проверяем, что категория присутствует в labels
        if data['labels']:
            assert product_category.name in data['labels']


# ==================== LowStockReportView ====================
@pytest.mark.django_db
class TestLowStockReportView:
    def test_low_stock_report_view_renders_correct_template(self, client, user):
        """Тест, что LowStockReportView использует правильный шаблон"""
        client.force_login(user)
        url = reverse('low_stock_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'reports/low_stock_report.html' in [t.name for t in response.templates]
    
    def test_low_stock_report_view_with_low_stock_materials(self, client, user, material):
        """Тест LowStockReportView с материалами с низким запасом"""
        # Настраиваем материал с низким запасом
        material.quantity = 5.00
        material.min_quantity = 10.00
        material.save()
        
        client.force_login(user)
        url = reverse('low_stock_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'materials' in response.context
        materials = response.context['materials']
        
        # Проверяем, что материал попал в отчет
        assert materials.count() >= 1
        material_in_report = materials.first()
        assert material_in_report.needed_quantity == 5.00  # 10 - 5
    
    def test_low_stock_report_view_without_low_stock(self, client, user, material):
        """Тест LowStockReportView без материалов с низким запасом"""
        # Настраиваем материал с достаточным запасом
        material.quantity = 20.00
        material.min_quantity = 10.00
        material.save()
        
        client.force_login(user)
        url = reverse('low_stock_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'materials' in response.context
        # Материал не должен попасть в отчет
        assert response.context['materials'].count() == 0
    
    def test_low_stock_report_view_ordering(self, client, user, material):
        """Тест сортировки в LowStockReportView"""
        # Создаем второй материал
        material2 = Material.objects.create(
            name="Второй материал",
            article="TEST-MAT-002",
            category=material.category,
            quantity=2.00,
            min_quantity=10.00,
            color=material.color,
            unit=material.unit
        )
        
        # Первый материал с меньшим количеством
        material.quantity = 1.00
        material.min_quantity = 10.00
        material.save()
        
        client.force_login(user)
        url = reverse('low_stock_report')
        response = client.get(url)
        
        assert response.status_code == 200
        materials = response.context['materials']
        
        # Проверяем сортировку по имени
        if materials.count() >= 2:
            names = [m.name for m in materials]
            assert names == sorted(names)



# ==================== StockAgeingReportView ====================
@pytest.mark.django_db
class TestStockAgeingReportView:
    def test_stock_ageing_report_view_renders_correct_template(self, client, user):
        """Тест, что StockAgeingReportView использует правильный шаблон"""
        client.force_login(user)
        url = reverse('stock_ageing_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'reports/stock_ageing_report.html' in [t.name for t in response.templates]
    
    def test_stock_ageing_report_view_context(self, client, user, product, material, product_operation_incoming, material_operation_incoming):
        """Тест контекста StockAgeingReportView"""
        client.force_login(user)
        url = reverse('stock_ageing_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'items' in response.context
        assert 'current_sort' in response.context
        assert response.context['current_sort'] == 'asc'
    
    
    def test_stock_ageing_report_view_pagination(self, client, user):
        """Тест пагинации в StockAgeingReportView"""
        client.force_login(user)
        url = reverse('stock_ageing_report')
        response = client.get(url)
        
        assert response.status_code == 200
        assert 'paginator' in response.context
        assert 'page_obj' in response.context
    
    def test_stock_ageing_report_view_excludes_zero_quantity(self, client, user, product):
        """Тест, что товары с нулевым количеством не попадают в отчет"""
        # Продукт с нулевым количеством
        product.total_quantity = 0
        product.save()
        
        client.force_login(user)
        url = reverse('stock_ageing_report')
        response = client.get(url)
        
        assert response.status_code == 200
        items = response.context['items']
        
        # Продукт с нулевым количеством не должен попасть в отчет
        product_in_report = any(item['sku'] == product.sku for item in items)
        assert not product_in_report

# ==================== Интеграционные тесты ====================
@pytest.mark.django_db
class TestReportsIntegration:
    def test_all_report_views_accessible(self, client, user):
        """Интеграционный тест: все отчеты доступны"""
        client.force_login(user)
        
        report_urls = [
            reverse('reports_home'),
            reverse('sales_over_time'),
            reverse('movement_report'),
            reverse('sales_by_product_report'),
            reverse('low_stock_report'),
            reverse('stock_ageing_report'),
        ]
        
        for url in report_urls:
            response = client.get(url)
            assert response.status_code == 200, f"URL {url} вернул статус {response.status_code}"
    
    def test_all_api_endpoints_accessible(self, client, user):
        """Интеграционный тест: все API эндпоинты доступны"""
        client.force_login(user)
        
        api_urls = [
            reverse('sales_chart_data_api'),
            reverse('sales_by_product_api'),
            reverse('sales_by_category_api'),
        ]
        
        for url in api_urls:
            response = client.get(url)
            assert response.status_code in [200, 400], f"URL {url} вернул статус {response.status_code}"


@pytest.mark.django_db
class TestReportsSecurity:
    def test_all_reports_redirect_anonymous(self, client):
        """Проверка, что анонимный пользователь перенаправляется на логин"""
        report_urls = [
            reverse('reports_home'),
            reverse('sales_over_time'),
            reverse('movement_report'),
            reverse('low_stock_report'),
            reverse('stock_ageing_report'),
        ]
        for url in report_urls:
            response = client.get(url)
            assert response.status_code == 302
            assert 'login/' in response.url


@pytest.mark.django_db
class TestSalesApiDeepDive:
    def test_sales_chart_api_filters_status(self, client, user, shipment, product):
        """API должен игнорировать заказы, которые не в статусе 'shipped'"""
        client.force_login(user)
        # Создаем ShipmentItem для заказа в статусе 'pending' (из фикстуры)
        ShipmentItem.objects.create(shipment=shipment, product=product, quantity=10, price=100)
        
        response = client.get(reverse('sales_chart_data_api'))
        data = json.loads(response.content)
        # Сумма должна быть 0, так как статус 'pending', а не 'shipped'
        assert sum(data['data']) == 0

    def test_sales_by_product_math(self, client, user, shipment, product):
        client.force_login(user)
        shipment.items.all().delete()
        
        # Если цена берется из продукта, меняем её в продукте
        product.price = Decimal('150.50')
        product.save()

        shipment.status = 'shipped'
        shipment.shipped_at = timezone.now()
        shipment.save()

        ShipmentItem.objects.create(shipment=shipment, product=product, quantity=10, price=product.price)
        
        response = client.get(reverse('sales_by_product_api'))
        data = json.loads(response.content)
        assert data['data'][0] == 1505.00


@pytest.mark.django_db
class TestLowStockEdgeCases:
    def test_low_stock_exact_boundary(self, client, user, material):
        """Материал должен попасть в отчет, если остаток РОВНО равен min_quantity"""
        client.force_login(user)
        material.quantity = Decimal('10.00')
        material.min_quantity = Decimal('10.00')
        material.save()
        
        response = client.get(reverse('low_stock_report'))
        assert material in response.context['materials']

    def test_low_stock_ignore_zero_threshold(self, client, user, material):
        """Если min_quantity = 0, материал не считается дефицитным, даже если остаток 0"""
        client.force_login(user)
        material.quantity = 0
        material.min_quantity = 0
        material.save()
        
        response = client.get(reverse('low_stock_report'))
        assert material not in response.context['materials']


@pytest.mark.django_db
class TestStockAgeingLogic:
    
    # Мокаем путь к задаче. 
    # ВАЖНО: Указываем путь к тому месту, где задача ВЫЗЫВАЕТСЯ (в сигналах)
    @patch('warehouse2.signals.update_stock_in_keycrm.delay')
    def test_stock_ageing_combined_sorting(self, mock_task, client, user, product, material):
        """Проверка ручной сортировки (Товары + Материалы)"""
        client.force_login(user)
        from django.utils import timezone
        from datetime import timedelta
        from django.contrib.contenttypes.models import ContentType

        # 1. Полная очистка
        ProductOperation.objects.all().delete()
        MaterialOperation.objects.all().delete()

        # 2. Фиксируем даты
        now = timezone.now().replace(microsecond=0)
        old_date = now - timedelta(days=10)
        new_date = now - timedelta(days=1)

        # 3. Создаем операцию для ТОВАРА
        product_ct = ContentType.objects.get_for_model(product.__class__)
        op_p = ProductOperation.objects.create(
            product=product,
            operation_type="incoming", # Проверьте, что в модели это строка или Enum
            quantity=1,
            user=user,
            content_type=product_ct,
            object_id=product.id
        )
        # Обновляем дату
        ProductOperation.objects.filter(id=op_p.id).update(timestamp=old_date)

        # 4. Создаем операцию для МАТЕРИАЛА
        op_m = MaterialOperation.objects.create(
            material=material,
            operation_type='incoming',
            quantity=1,
            user=user
        )
        MaterialOperation.objects.filter(id=op_m.id).update(date=new_date)

        # 5. Выполняем запрос
        res_asc = client.get(reverse('stock_ageing_report'), {'sort': 'asc'})
        items_asc = res_asc.context['items']

        # Проверки
        assert items_asc[0]['name'] == product.name
        assert items_asc[0]['age_days'] >= 10
        
        # Проверка, что мок сработал (задача вызывалась, но не выполнялась)
        assert mock_task.called

@pytest.mark.django_db
class TestMovementReportRobustness:
    def test_per_page_validation(self, client, user):
        """Проверка обработки некорректных параметров пагинации"""
        client.force_login(user)
        url = reverse('movement_report')
        
        # Кейс 1: Слишком большое число (должно ограничиться 500)
        response = client.get(url, {'per_page': 9999})
        assert response.context['per_page'] == 500
        
        # Кейс 2: Не число (должно сброситься на 25)
        response = client.get(url, {'per_page': 'invalid'})
        assert response.context['per_page'] == 25